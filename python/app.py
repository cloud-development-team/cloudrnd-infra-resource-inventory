from flask import Flask, render_template, jsonify, request, send_file
import boto3
import re
import os
import json
import subprocess
from datetime import datetime, timezone
from tempfile import NamedTemporaryFile
import pandas as pd
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from modules.vpc import list_vpcs
from modules.subnet import list_subnets
from modules.sg import list_security_groups
from modules.nacl import list_nacls
from modules.ec2 import list_ec2_instances
from modules.asg import list_auto_scaling_groups
from modules.elb import list_elbs
from modules.tg import list_target_groups
from modules.cloudfront import list_cloudfront_distributions
from modules.s3 import list_s3_buckets
from modules.iamrole import list_iam_roles
from modules.db import list_db_clusters
from modules.elasticache import list_elasticache_clusters
from modules.msk import list_kafka_clusters
from modules.sg_resource_mapper import map_sg_to_resources
from modules.sg_centric_rules import map_sg_rules_with_resources
from modules.sg_summary import map_sg_summary
from modules.route53 import list_route53_zones, list_zone_record_sets, sanitize_sheet_name
from openpyxl.worksheet.table import Table, TableStyleInfo

app = Flask(__name__)

RESOURCE_MAP = {
    "vpcs": list_vpcs,
    "subnets": list_subnets,
    "security-groups": list_security_groups,
    "nacl": list_nacls,
    "ec2": list_ec2_instances,
    "asg": list_auto_scaling_groups,
    "elbs": list_elbs,
    "target-groups": list_target_groups,
    "cloudfront": list_cloudfront_distributions,
    "s3": list_s3_buckets,
    "iam-roles": list_iam_roles,
    "database": list_db_clusters,
    "elasticache": list_elasticache_clusters,
    "msk": list_kafka_clusters
}

def get_aws_profiles(config_path="~/.aws/config"):
    profiles = []
    path = os.path.expanduser(config_path)
    if os.path.exists(path):
        with open(path, 'r') as f:
            content = f.read()
            profiles = re.findall(r'\[profile\s+([^\]]+)\]', content)
    return profiles

def has_valid_sso_token():
    cache_dir = os.path.expanduser("~/.aws/sso/cache")
    if not os.path.isdir(cache_dir):
        return False

    for filename in os.listdir(cache_dir):
        file_path = os.path.join(cache_dir, filename)
        try:
            with open(file_path, 'r') as f:
                data = json.load(f)

                # expiresAt이 없으면 스킵
                if "expiresAt" not in data:
                    continue

                # 시간 비교 (UTC)
                expires = datetime.fromisoformat(data["expiresAt"].replace("Z", "+00:00"))
                if expires > datetime.now(timezone.utc):
                    return True  # 유효한 SSO 토큰 있음
        except Exception as e:
            continue

    return False  # 유효한 토큰 없음

@app.route('/')
def index():
    profiles = get_aws_profiles()
    sso_required = not has_valid_sso_token()
    return render_template('index.html', resource_keys=list(RESOURCE_MAP.keys()), profile_list=profiles, sso_login_required=sso_required)

@app.route('/sso-login', methods=['POST'])
def sso_login():
    try:
        proc = subprocess.Popen(
            # ["aws", "sso", "login", "--sso-session", "hanwhavision", "--use-device"],
            ["aws", "sso", "login", "--sso-session", "hanwhavision"],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True
        )

        login_url = None
        user_code = None

        for line in iter(proc.stdout.readline, ''):
            if "https://" in line and "device" in line:
                login_url = re.search(r"(https://[^\s]+)", line)
                if login_url:
                    login_url = login_url.group(1)
            if re.match(r"^[A-Z0-9]{4}-[A-Z0-9]{4}$", line.strip()):
                user_code = line.strip()
            if login_url and user_code:
                break

        return jsonify({
            "login_url": login_url,
            "user_code": user_code
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/check-sso')
def check_sso():
    return jsonify({"ok": has_sso_token("hanwhavision")})

@app.route('/logout-sso', methods=['POST'])
def logout_sso():
    try:
        cache_dir = os.path.expanduser("~/.aws/sso/cache")
        if os.path.isdir(cache_dir):
            for filename in os.listdir(cache_dir):
                file_path = os.path.join(cache_dir, filename)
                try:
                    os.remove(file_path)
                except Exception as e:
                    print(f"Failed to delete {file_path}: {e}")
        return jsonify({"message": "SSO token cleared"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/<resource>')
def get_resource(resource):
    if resource not in RESOURCE_MAP:
        return jsonify({"error": "Unsupported resource"}), 404

    profile = request.args.get("profile", "sightmind-prod")
    try:
        session = boto3.Session(profile_name=profile)
        result = RESOURCE_MAP[resource](session)
        if not result:
            return jsonify({"columns": [], "rows": []})
        
        columns = list(result[0].keys())
        rows = [list(item.values()) for item in result]
        return jsonify({"columns": columns, "rows": rows})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def save_excel_with_format(sheet_dict, filename):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for sheet_name, df in sheet_dict.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    wb = load_workbook(filename)
    header_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0) + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 필터 및 테이블 스타일 추가
        if ws.max_row > 1 and ws.max_column > 0:
            table_range = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            table = Table(displayName="Table1", ref=table_range)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=False, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)

    wb.save(filename)

@app.route('/download/overall')
def download_overall():
    profile = request.args.get("profile")
    if not profile:
        return "Profile is required", 400

    try:
        session = boto3.Session(profile_name=profile)
        file_name = f"{profile}_overall_inventory_{datetime.now().strftime('%y_%m_%d')}.xlsx"

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            sheet_dict = {}
            for key, func in RESOURCE_MAP.items():
                try:
                    data = func(session)
                    if data:
                        sheet_dict[key] = pd.DataFrame(data)
                except Exception as e:
                    print(f"Skipping {key}: {e}")

            save_excel_with_format(sheet_dict, tmp.name)
            return send_file(tmp.name, download_name=file_name, as_attachment=True)

    except Exception as e:
        return str(e), 500

@app.route('/download/route53')
def download_route53_detail():
    profile = request.args.get("profile")
    if not profile:
        return "Profile is required", 400

    try:
        session = boto3.Session(profile_name=profile)
        file_name = f"{profile}_route53_detail_inventory_{datetime.now().strftime('%y_%m_%d')}.xlsx"

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            zones, summary_df = list_route53_zones(session)
            sheet_dict = {"Hosted Zones": summary_df}
            for zone in zones:
                zone_name = zone['Name'].rstrip('.')
                sheet_name = sanitize_sheet_name(zone_name)
                records_df = list_zone_record_sets(session, zone['Id'])
                sheet_dict[sheet_name] = records_df

            save_excel_with_format(sheet_dict, tmp.name)
            return send_file(tmp.name, download_name=file_name, as_attachment=True)

    except Exception as e:
        return str(e), 500

@app.route('/download/sg')
def download_sg_detail():
    profile = request.args.get("profile")
    if not profile:
        return "Profile is required", 400

    try:
        session = boto3.Session(profile_name=profile)
        file_name = f"{profile}_sg_detail_inventory_{datetime.now().strftime('%y_%m_%d')}.xlsx"

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            sheet_dict = {
                "SG-Resource No Rules": pd.DataFrame(map_sg_summary(session)),
                "SG-Resource Rules": pd.DataFrame(map_sg_rules_with_resources(session)),
                "Resource-SG-Rules": pd.DataFrame(map_sg_to_resources(session))
            }
            save_excel_with_format(sheet_dict, tmp.name)
            return send_file(tmp.name, download_name=file_name, as_attachment=True)

    except Exception as e:
        return str(e), 500

@app.route('/download/sg_report')
def download_sg_report():
    profile = request.args.get("profile")
    if not profile:
        return "Profile is required", 400

    try:
        session = boto3.Session(profile_name=profile)
        sg_data = map_sg_rules_with_resources(session)
        if not sg_data:
            return "No SG data available", 404

        df = pd.DataFrame(sg_data)

        def analyze_sg(row):
            findings = []
            direction = row.get("Direction", "")
            port = str(row.get("Port Range", ""))
            source = str(row.get("Src Origin", ""))
            source_name = str(row.get("Src Parsed", ""))
            destination = str(row.get("Des Origin", ""))
            destination_name = str(row.get("Des Parsed", ""))
            protocol = str(row.get("Protocol", "")).lower()
            sg_id = row.get("Security Group ID", "")

            is_sg_source = source.startswith("sg-")
            is_sg_dest = destination.startswith("sg-")

            
            # Rule 1: overly open to 0.0.0.0/0
            if direction == "Inbound" and source == "0.0.0.0/0" and (
                re.search(r"^22$|^22[-:]", port) or protocol == "all"
            ):
                findings.append("Inbound 0.0.0.0/0 open (22/ALL)")

            if direction == "Outbound" and destination == "0.0.0.0/0" and protocol == "all":
                findings.append("Outbound 0.0.0.0/0 open (ALL)")

            # Rule 2: unused SG
            if str(row.get("Usage", "")).strip().upper() == "FALSE":
                findings.append("Unused SG")

            # Rule 3: SG reference mismatch
            # --- Outbound: referencing a destination SG ---
            if direction == "Outbound" and is_sg_dest:
                matched_inbound = (
                    (df["Security Group ID"] == destination) &
                    (df["Direction"] == "Inbound") &
                    (df["Src Origin"] == sg_id)
                )
                if not matched_inbound.any():
                    # Check if target SG is just open to 0.0.0.0/0
                    open_inbound = (
                        (df["Security Group ID"] == destination) &
                        (df["Direction"] == "Inbound") &
                        (df["Src Origin"] == "0.0.0.0/0")
                    )
                    if open_inbound.any():
                        findings.append(f"Outbound references {destination_name} but no matching inbound (note: {destination_name} open to 0.0.0.0/0)")
                    else:
                        findings.append(f"Outbound references {destination_name} but no matching inbound")

            # --- Inbound: referencing a source SG ---
            if direction == "Inbound" and is_sg_source:
                matched_outbound = (
                    (df["Security Group ID"] == source) &
                    (df["Direction"] == "Outbound") &
                    (df["Des Origin"] == sg_id)
                )
                if not matched_outbound.any():
                    # Check if source SG is open to 0.0.0.0/0
                    open_outbound = (
                        (df["Security Group ID"] == source) &
                        (df["Direction"] == "Outbound") &
                        (df["Des Origin"] == "0.0.0.0/0")
                    )
                    if open_outbound.any():
                        findings.append(f"Inbound references {source_name} but no matching outbound (note: {source_name} open to 0.0.0.0/0)")
                    else:
                        findings.append(f"Inbound references {source_name} but no matching outbound")

            return ", ".join(findings)


        df["Findings"] = df.apply(analyze_sg, axis=1)
        df_filtered = df[df["Findings"] != ""]
        columns_to_drop = [
            "Usage", "Region", "Src Origin", "Des Origin", 
            "Resource Name", "Resource ID", "Resource Type", 
            "ENI ID", "Private IP", "Security Group ID", "SG Description"
        ]
        df_filtered = df_filtered.drop(columns=[col for col in columns_to_drop if col in df_filtered.columns])
        df_filtered = df_filtered.drop_duplicates()
        cols = df_filtered.columns.tolist()
        if "Findings" in cols:
            cols.insert(0, cols.pop(cols.index("Findings")))
            df_filtered = df_filtered[cols]

        file_name = f"{profile}_sg_findings_{datetime.now().strftime('%y_%m_%d')}.xlsx"
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            sheet_dict = {"SG Findings": df_filtered}
            save_excel_with_format(sheet_dict, tmp.name)
            return send_file(tmp.name, download_name=file_name, as_attachment=True)

    except Exception as e:
        return str(e), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)